# Importações
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import openpyxl
from PIL import Image, ImageDraw, ImageFont

# Função para selecionar o arquivo .xlsx
def select_excel_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        excel_entry.delete(0, "end")
        excel_entry.insert(0, filename)

# Função para selecionar o local e o nome do arquivo .png
def select_folder():
    foldername = filedialog.askdirectory()
    if foldername:
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, foldername)

# Função para exibir a mensagem de conclusão
def show_completion_message():
    messagebox.showinfo("Concluído", "Todos os certificados foram gerados com sucesso!")


# Função para processar o arquivo .xlsx e gerar os arquivos .png
def process_files(): 
    excel_filename = excel_entry.get()
    folder_path = folder_entry.get()

    if excel_filename and folder_path:
        try:
            workbook = openpyxl.load_workbook(excel_filename)
            sheet = workbook.active
            total_rows = sheet.max_row - 1  # Excluindo a primeira linha de cabeçalho
            progress_bar["maximum"] = total_rows

            # Inicializar variáveis de iteração
            global indice
            indice = 1

            # Agendar a execução da função de processamento em intervalos regulares
            root.after(100, process_next_row, sheet, folder_path, total_rows)

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")

# Função para processar a próxima linha do arquivo .xlsx
def process_next_row(sheet, folder_path, total_rows):
    global indice

    if indice <= total_rows:
        try:
            linha = next(sheet.iter_rows(min_row=indice+1))  # Avançar para a próxima linha
            nome_curso = linha[0].value
            nome_participante = linha[1].value
            tipo_participacao = linha[2].value
            carga_horaria = linha[5].value
            data_inicio = linha[3].value
            data_fim = linha[4].value
            data_emissao = linha[6].value

            fonte_nome = ImageFont.truetype('./fontes/tahomabd.ttf', 90)
            fonte_geral = ImageFont.truetype('./fontes/tahoma.ttf', 80)
            fonte_data = ImageFont.truetype('./fontes/tahoma.ttf', 55)

            image = Image.open('./certificado_padrao.jpg')
            draw = ImageDraw.Draw(image)
            draw.text((1015, 827), nome_participante, fill='black', font=fonte_nome)
            draw.text((1060, 950), nome_curso, fill='black', font=fonte_geral)
            draw.text((1425, 1070), tipo_participacao, fill='black', font=fonte_geral)
            draw.text((1480, 1190), str(carga_horaria), fill='black', font=fonte_geral)
            draw.text((750, 1785), str(data_inicio), fill='black', font=fonte_data)
            draw.text((750, 1930), str(data_fim), fill='black', font=fonte_data)
            draw.text((2230, 1930), str(data_emissao), fill='black', font=fonte_data)

            image.save(f'{folder_path}/{indice} {nome_participante} certificado.png')

            # Atualizar a barra de progresso e rótulo
            progress_bar["value"] = indice
            progress_label.configure(text=f"Progresso: {indice}/{total_rows}")

            # Agendar a próxima execução da função para a próxima linha
            indice += 1
            root.after(100, process_next_row, sheet, folder_path, total_rows)

        except StopIteration:
            # O fim do arquivo foi alcançado, exibir mensagem de conclusão
            show_completion_message("Concluído", "Certificados gerados com sucesso!")
    else:
        # Todas as linhas foram processadas, não há mais nada a fazer
        pass
        show_completion_message()

# Criar a janela principal
root = ctk.CTk()
root.title("Gerador de Certificados")
root.geometry("600x300")

# Rótulos e entradas para selecionar o arquivo .xlsx e o local para salvar os arquivos .png
excel_label = ctk.CTkLabel(root, text="Selecione o arquivo .xlsx :")
excel_label.grid(row=0, column=0, padx=5, pady=5)
excel_entry = ctk.CTkEntry(root, width=50)
excel_entry.grid(row=0, column=1, padx=5, pady=5)
excel_button = ctk.CTkButton(root, text="Selecionar", command=select_excel_file)
excel_button.grid(row=0, column=2, padx=5, pady=5)

folder_label = ctk.CTkLabel(root, text="Selecione a pasta para salvar os certificados:")
folder_label.grid(row=1, column=0, padx=5, pady=5)
folder_entry = ctk.CTkEntry(root, width=50)
folder_entry.grid(row=1, column=1, padx=5, pady=5)
folder_button = ctk.CTkButton(root, text="Selecionar", command=select_folder)
folder_button.grid(row=1, column=2, padx=5, pady=5)

# Botão para iniciar o processamento
process_button = ctk.CTkButton(root, text="Gerar Certificados", command=process_files)
process_button.grid(row=2, column=1, padx=5, pady=5)

# Barra de progresso
progress_label = ctk.CTkLabel(root, text="Progresso: 0/0")
progress_label.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.grid(row=4, column=0, columnspan=3, padx=5, pady=5)

# Iniciar a interface gráfica
root.mainloop()
